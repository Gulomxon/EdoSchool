import json
from django.contrib import admin
from .forms import AccountForm
from .models import Account
from django.contrib.auth.models import Group
from django.utils.html import format_html
from image_cropping import ImageCropWidget
from django.db import models
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# admin.site.register(Account)
admin.site.unregister(Group)
admin.site.site_header = "EdoSchool"

@admin.register(Account)
class AccountAdmin(admin.ModelAdmin):
    form = AccountForm
    list_display = ("username","first_name", "last_name", "phone", "profile_image")
    list_display_links = ("username","first_name", "last_name")
    list_per_page = 10
    list_editable = ("phone",)
    change_form_template = 'admin/myapp/my_model_change_form.html'
    actions = ['export_to_excel', 'export_as_json', 'export_to_json']
    
    search_fields = ("first_name", "last_name", "phone")
    
# show image in admin page 
    def profile_image(self, obj):
        if obj.profile:
            return format_html('<img src="{}" height="100px" />'.format(obj.profile.url))
        else:
            return ''

    profile_image.short_description = 'Profile Image'
    
# add ections
    
    def export_to_excel(modeladmin, request, queryset):

        wb = Workbook()

        ws = wb.active
        ws.title = 'account'
        headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Birthday']
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws['{}1'.format(col_letter)]
            cell.value = header_title
            cell.font = Font(bold=True, size=18, name="Times New Roman")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 30
        ws.column_dimensions[get_column_letter(4)].width = 20
        ws.column_dimensions[get_column_letter(5)].width = 20
        
        for row_num, obj in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=obj.first_name).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_num, column=2, value=obj.last_name).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_num, column=3, value=obj.email).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_num, column=4, value=str(obj.phone)).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_num, column=5, value=obj.birthday).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=accounts.xlsx'

        wb.save(response)

        return response
    export_to_excel.short_description = "Export to Excel"
    
    def export_to_json(self, request, queryset):
        data = []
        for obj in queryset:
            data.append({
                'first_name': obj.first_name,
                'last_name': obj.last_name,
                'email': obj.email,
                'phone': str(obj.phone),
                'birthday': obj.birthday.strftime('%Y-%m-%d'),
            })
        response = HttpResponse(json.dumps(data, indent=4), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="mydata.json"'
        return response
    export_to_json.short_description = 'Export to JSON'
    
    
    def export_as_json(self, request, queryset):
        data = []
        for obj in queryset:
            data.append({
                'first_name': obj.first_name,
                'last_name': obj.last_name,
                'email': obj.email,
                'phone': str(obj.phone),
                'birthday': obj.birthday.strftime('%Y-%m-%d'),
            })
        return JsonResponse(data, json_dumps_params={'indent': 4}, safe=False)
    export_as_json.short_description = 'Show in JSON format'